import os
import openpyxl
from tqdm import tqdm

# ============================================================
# SETTINGS
# ============================================================

# Excel file containing the numbers to search
input_file = r"C:\Users\USER\Downloads\Sem 5 DONE\New folder (2)\nrc.xlsx"

# Folder containing the Excel files to search
search_folder = r"C:\Users\USER\Downloads\Sem 5 DONE\New folder (2)\PMEC"

# Output file
output_file = r"C:\Users\USER\Downloads\Sem 5 DONE\New folder (2)\PMEC output.xlsx"


# ============================================================
# STEP 1: READ NUMBERS FROM COLUMN A
# ============================================================

print("Reading numbers from input file...")

wb_input = openpyxl.load_workbook(input_file, data_only=True)
ws_input = wb_input.active

numbers_to_find = set()

for row in ws_input.iter_rows(min_col=1, max_col=1):
    value = row[0].value

    if value is not None:
        value = str(value).strip()

        if value:
            numbers_to_find.add(value)

wb_input.close()

print(f"Numbers to search: {len(numbers_to_find)}")


# ============================================================
# STEP 2: FIND ALL EXCEL FILES INCLUDING SUBFOLDERS
# ============================================================

excel_extensions = (".xlsx", ".xlsm", ".xltx", ".xltm")

excel_files = []

for root, dirs, files in os.walk(search_folder):

    for file_name in files:

        if file_name.lower().endswith(excel_extensions):

            file_path = os.path.join(root, file_name)

            if os.path.abspath(file_path) != os.path.abspath(output_file):
                excel_files.append(file_path)


print(f"Excel files found: {len(excel_files)}")


# ============================================================
# STEP 3: CREATE OUTPUT WORKBOOK
# ============================================================

wb_output = openpyxl.Workbook()
ws_output = wb_output.active
ws_output.title = "Search Results"

# These are the first columns in the output
ws_output.append([
    "Found Number",
    "File Name",
    "Sheet Name",
    "Original Row Number"
])

total_matches = 0


# ============================================================
# STEP 4: SEARCH ALL EXCEL FILES
# ============================================================

for file_path in tqdm(
    excel_files,
    desc="Searching Excel files",
    unit="file"
):

    file_name = os.path.basename(file_path)

    try:

        keep_vba = file_name.lower().endswith(".xlsm")

        wb = openpyxl.load_workbook(
            file_path,
            read_only=True,
            data_only=True,
            keep_vba=keep_vba
        )

        # Search every sheet
        for ws in wb.worksheets:

            for row in ws.iter_rows():

                # Check every cell in the row
                found_number = None

                for cell in row:

                    if cell.value is None:
                        continue

                    cell_value = str(cell.value).strip()

                    if cell_value in numbers_to_find:
                        found_number = cell_value
                        break

                # ====================================================
                # NUMBER FOUND - COPY THE COMPLETE ROW
                # ====================================================

                if found_number is not None:

                    # Basic information
                    output_row = [
                        found_number,
                        file_name,
                        ws.title,
                        row[0].row
                    ]

                    # Add every cell value from the original row
                    for cell in row:
                        output_row.append(cell.value)

                    ws_output.append(output_row)

                    total_matches += 1

        wb.close()

    except Exception as e:

        tqdm.write(f"ERROR: {file_path}")
        tqdm.write(str(e))


# ============================================================
# STEP 5: SAVE OUTPUT
# ============================================================

wb_output.save(output_file)


# ============================================================
# COMPLETED
# ============================================================

print("\n======================================")
print("SEARCH COMPLETED")
print("======================================")

print(f"Excel files searched : {len(excel_files)}")
print(f"Matches found        : {total_matches}")
print(f"Results saved to     : {output_file}")
